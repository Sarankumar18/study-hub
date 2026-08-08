"""
Gmail Job Search Email Reader

Reads job-related emails from Gmail and writes a summary to
career/gmail_job_updates.md for the AI mentor to read.

Usage:
  python3 career/gmail_job_reader.py

First run will open browser for Google OAuth authorization.
Subsequent runs use cached token.
"""

import os
import json
import base64
import re
from datetime import datetime, timedelta
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

PROJECT_ROOT = Path(__file__).resolve().parent.parent
CREDS_FILE = PROJECT_ROOT / '.secrets' / 'gmail_credentials.json'
TOKEN_FILE = PROJECT_ROOT / '.secrets' / 'token.json'
OUTPUT_FILE = PROJECT_ROOT / 'career' / 'gmail_job_updates.md'
LAST_RUN_FILE = PROJECT_ROOT / '.secrets' / 'gmail_last_run.json'

JOB_KEYWORDS = [
    'application', 'interview', 'assessment', 'coding challenge',
    'offer', 'rejected', 'shortlisted', 'next steps', 'hiring',
    'recruiter', 'talent', 'opportunity', 'position', 'role',
    'onsite', 'phone screen', 'technical round', 'HR round',
    'congratulations', 'unfortunately', 'move forward',
    'schedule', 'availability', 'compensation', 'package',
]

COMPANY_KEYWORDS = [
    'google', 'meta', 'stripe', 'uber', 'databricks', 'confluent',
    'snowflake', 'microsoft', 'amazon', 'mastercard', 'adyen',
    'booking.com', 'booking', 'atlassian', 'spotify', 'zalando',
    'servicenow', 'adobe', 'paypal', 'klarna', 'wise', 'elastic',
    'mongodb', 'cloudflare', 'shopee', 'sea labs', 'grab',
    'airbnb', 'oracle', 'cisco', 'sap', 'intuit', 'ebay',
    'rubrik', 'nutanix', 'n26', 'asml', 'tomtom', 'philips',
    'naukri', 'linkedin', 'hired', 'lever', 'greenhouse',
    'workday', 'smartrecruiters', 'ashby',
]


def get_installed_creds_file():
    """Convert 'web' type credentials to 'installed' type if needed."""
    with open(CREDS_FILE) as f:
        data = json.load(f)

    if 'installed' in data:
        return str(CREDS_FILE)

    if 'web' in data:
        installed_file = CREDS_FILE.parent / 'gmail_credentials_installed.json'
        installed_data = {'installed': data['web']}
        with open(installed_file, 'w') as f:
            json.dump(installed_data, f)
        return str(installed_file)

    raise ValueError('Unrecognized credential format')


def authenticate():
    creds = None

    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            creds_file = get_installed_creds_file()
            flow = InstalledAppFlow.from_client_secrets_file(creds_file, SCOPES)
            creds = flow.run_local_server(port=0)

        TOKEN_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(TOKEN_FILE, 'w') as f:
            f.write(creds.to_json())

    return creds


def get_header(headers, name):
    for h in headers:
        if h['name'].lower() == name.lower():
            return h['value']
    return ''


def is_job_related(subject, sender, snippet):
    text = f"{subject} {sender} {snippet}".lower()
    return any(kw in text for kw in JOB_KEYWORDS) or any(kw in text for kw in COMPANY_KEYWORDS)


def extract_body_text(payload):
    if payload.get('body', {}).get('data'):
        return base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8', errors='ignore')

    parts = payload.get('parts', [])
    for part in parts:
        if part['mimeType'] == 'text/plain' and part.get('body', {}).get('data'):
            return base64.urlsafe_b64decode(part['body']['data']).decode('utf-8', errors='ignore')

    for part in parts:
        result = extract_body_text(part)
        if result:
            return result

    return ''


def clean_body(body, max_chars=500):
    body = re.sub(r'<[^>]+>', ' ', body)
    body = re.sub(r'\s+', ' ', body).strip()
    if len(body) > max_chars:
        body = body[:max_chars] + '...'
    return body


def get_last_run_date():
    if LAST_RUN_FILE.exists():
        with open(LAST_RUN_FILE) as f:
            data = json.load(f)
            return data.get('last_run', None)
    return None


def save_last_run_date():
    LAST_RUN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(LAST_RUN_FILE, 'w') as f:
        json.dump({'last_run': datetime.now().strftime('%Y/%m/%d'), 'last_run_iso': datetime.now().isoformat()}, f)


def fetch_job_emails(days=None, max_results=100):
    creds = authenticate()
    service = build('gmail', 'v1', credentials=creds)

    if days is not None:
        after_date = (datetime.now() - timedelta(days=days)).strftime('%Y/%m/%d')
    else:
        last_run = get_last_run_date()
        if last_run:
            after_date = last_run
        else:
            after_date = (datetime.now() - timedelta(days=7)).strftime('%Y/%m/%d')

    query = f'after:{after_date}'

    results = service.users().messages().list(
        userId='me', q=query, maxResults=max_results
    ).execute()

    messages = results.get('messages', [])
    job_emails = []

    for msg_meta in messages:
        msg = service.users().messages().get(
            userId='me', id=msg_meta['id'], format='full'
        ).execute()

        headers = msg.get('payload', {}).get('headers', [])
        subject = get_header(headers, 'Subject')
        sender = get_header(headers, 'From')
        date = get_header(headers, 'Date')
        snippet = msg.get('snippet', '')
        labels = msg.get('labelIds', [])

        if is_noise(sender, subject):
            continue

        if not is_job_related(subject, sender, snippet):
            continue

        body = extract_body_text(msg.get('payload', {}))
        body_clean = clean_body(body, max_chars=800)

        gmail_section = 'Primary'
        if 'CATEGORY_PROMOTIONS' in labels:
            gmail_section = 'Promotions'
        elif 'CATEGORY_SOCIAL' in labels:
            gmail_section = 'Social'
        elif 'CATEGORY_UPDATES' in labels:
            gmail_section = 'Updates'
        elif 'CATEGORY_FORUMS' in labels:
            gmail_section = 'Forums'
        elif 'SPAM' in labels:
            gmail_section = 'Spam'

        company = extract_company_from_email(sender, subject)

        job_emails.append({
            'date': date,
            'from': sender,
            'subject': subject,
            'snippet': snippet,
            'body_preview': body_clean,
            'gmail_section': gmail_section,
            'company': company,
            'labels': labels,
        })

    return job_emails


def write_output(emails):
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    last_run = get_last_run_date() or 'first run'

    lines = [
        f'# Gmail Job Updates',
        f'',
        f'> Last fetched: {now}',
        f'> Since: {last_run}',
        f'> Job-related found: {len(emails)}',
        f'',
        f'---',
        f'',
    ]

    if not emails:
        lines.append('No job-related emails found in the last 7 days.')
    else:
        for i, email in enumerate(emails, 1):
            lines.append(f'## {i}. {email["subject"]}')
            lines.append(f'')
            lines.append(f'**From:** {email["from"]}')
            lines.append(f'**Date:** {email["date"]}')
            lines.append(f'')
            lines.append(f'**Preview:**')
            lines.append(f'```')
            lines.append(email['body_preview'] or email['snippet'])
            lines.append(f'```')
            lines.append(f'')
            lines.append(f'---')
            lines.append(f'')

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text('\n'.join(lines))
    print(f'Wrote {len(emails)} job emails to {OUTPUT_FILE}')


NOISE_SENDERS = [
    'mailer-daemon', 'no-reply@accounts.google.com',
    'cloudplatform-noreply@google.com', 'payments-noreply@google.com',
    'platformnotifications-noreply@google.com', 'googlecloud@google.com',
    'noreply@online.proteantech.in', 'groww', 'zomato', 'swiggy',
    'admission', 'srmist.edu', 'codingchallenges@substack',
]


def is_noise(sender, subject):
    text = f"{sender} {subject}".lower()
    return any(n in text for n in NOISE_SENDERS)


def extract_company_from_email(sender, subject):
    for company in COMPANY_KEYWORDS:
        if company.lower() in sender.lower() or company.lower() in subject.lower():
            return company.title()
    sender_domain = re.search(r'@([a-zA-Z0-9.-]+)', sender)
    if sender_domain:
        domain = sender_domain.group(1).split('.')[0]
        skip = {'gmail', 'google', 'yahoo', 'outlook', 'hotmail'}
        if domain not in skip:
            return domain.title()
    return 'Unknown'


def write_excel(emails):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Skipping Excel export.")
        return

    wb = Workbook()

    ws = wb.active
    ws.title = "All Job Emails"

    headers = ['Date', 'Gmail Section', 'Company', 'From', 'Subject', 'Preview']
    header_fill = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side("thin", color="CCCCCC"),
        right=Side("thin", color="CCCCCC"),
        top=Side("thin", color="CCCCCC"),
        bottom=Side("thin", color="CCCCCC"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_num, email in enumerate(emails, 2):
        values = [
            email['date'][:25],
            email.get('gmail_section', 'Unknown'),
            email.get('company', 'Unknown'),
            email['from'][:60],
            email['subject'][:100],
            (email.get('body_preview', '') or email['snippet'])[:300],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 80

    excel_path = PROJECT_ROOT / 'career' / 'job_tracker.xlsx'
    wb.save(str(excel_path))
    print(f'Wrote Excel tracker to {excel_path}')


if __name__ == '__main__':
    import sys

    days = None
    if '--days' in sys.argv:
        idx = sys.argv.index('--days')
        if idx + 1 < len(sys.argv):
            days = int(sys.argv[idx + 1])

    if '--all' in sys.argv:
        days = 30

    emails = fetch_job_emails(days=days, max_results=200)
    write_output(emails)
    write_excel(emails)
    save_last_run_date()
    print(f'Next run will fetch emails from {datetime.now().strftime("%Y/%m/%d")} onwards.')
