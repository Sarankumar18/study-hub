"""
Application Tracker Builder

Builds a well-designed Excel application tracker from Gmail data.
Run after gmail_job_reader.py to create/update the tracker.

Usage:
  python3 career/build_tracker.py
"""

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TRACKER_FILE = PROJECT_ROOT / 'career' / 'application_tracker.xlsx'

# Colors
C_HEADER = "1a73e8"
C_HEADER2 = "34a853"
C_HEADER3 = "ea4335"
C_HEADER4 = "f59e0b"
C_LIGHT_BLUE = "e8f0fe"
C_LIGHT_GREEN = "e6f4ea"
C_LIGHT_RED = "fce8e6"
C_LIGHT_YELLOW = "fef7e0"
C_WHITE = "FFFFFF"
C_DARK = "1f1f1f"

HEADER_FONT = Font(bold=True, color=C_WHITE, size=11, name="Calibri")
BODY_FONT = Font(size=11, name="Calibri", color=C_DARK)
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color=C_HEADER)

BORDER = Border(
    left=Side("thin", color="CCCCCC"),
    right=Side("thin", color="CCCCCC"),
    top=Side("thin", color="CCCCCC"),
    bottom=Side("thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, num_cols, color):
    fill = PatternFill("solid", fgColor=color)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER


def style_row(ws, row, num_cols, alt=False):
    fill = PatternFill("solid", fgColor=C_LIGHT_BLUE if alt else C_WHITE)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = LEFT if col in (2, 3, 4, 16) else CENTER
        cell.border = BORDER


def add_validation(ws, col_letter, start_row, end_row, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select from dropdown"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def build_applications_sheet(wb):
    ws = wb.active
    ws.title = "Applications"

    headers = [
        "#",
        "Company",
        "Position",
        "Location",
        "Country",
        "Priority",
        "Match Score",
        "Applied Date",
        "Source",
        "Application Status",
        "Callback",
        "Interview Round",
        "Interview Date",
        "Follow-up Date",
        "Follow-up Done",
        "Notes",
    ]

    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = "APPLICATION TRACKER"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    ws.merge_cells("A2:P2")
    subtitle = ws["A2"]
    subtitle.value = f"Senior Backend Engineer | Java | International | Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    subtitle.font = Font(size=10, color="666666", name="Calibri")
    subtitle.alignment = CENTER

    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header(ws, 4, len(headers), C_HEADER)

    # Pre-fill with known applications
    applications = [
        [1, "Stripe", "Backend Engineer/API, Payments & Risk", "Dublin", "Ireland", "S", "9/10", "2026-08-08", "Career Page", "Applied", "No", "—", "—", "2026-08-15", "No", ""],
        [2, "Stripe", "SWE, Payments, Risk & Premium Merchant", "Dublin", "Ireland", "S", "8/10", "2026-08-08", "Career Page", "Applied", "No", "—", "—", "2026-08-15", "No", ""],
        [3, "Sea Labs Indonesia", "Back End Engineer, Marketplace Order Ops", "Jakarta", "Indonesia", "A", "7/10", "2026-08-08", "Career Page", "Applied", "No", "—", "—", "2026-08-15", "No", "Cover letter sent"],
        [4, "Sea", "Backend Engineer", "Singapore", "Singapore", "A", "7/10", "2026-08-08", "Career Page", "Applied", "No", "—", "—", "2026-08-15", "No", ""],
        [5, "Sea", "Senior Backend Engineer", "Singapore", "Singapore", "A", "8/10", "2026-08-08", "Career Page", "Applied", "No", "—", "—", "2026-08-15", "No", ""],
        [6, "Yi Connect", "Software Development", "Dubai", "UAE", "B", "5/10", "2026-08-08", "LinkedIn", "Applied", "No", "—", "—", "—", "No", ""],
        [7, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [8, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [9, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [10, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [11, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [12, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [13, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [14, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [15, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [16, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [17, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [18, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [19, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [20, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]

    for row_idx, app in enumerate(applications, 5):
        for col_idx, value in enumerate(app, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_row(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))

    # Dropdowns
    end_row = 4 + len(applications)
    add_validation(ws, "F", 5, end_row, ["S", "A", "B", "C"])
    add_validation(ws, "I", 5, end_row, ["Career Page", "LinkedIn", "Naukri", "NaukriGulf", "Foundit", "Referral", "Recruiter", "Other"])
    add_validation(ws, "J", 5, end_row, ["Not Applied", "Applied", "OA Received", "Phone Screen", "Technical Round", "System Design", "Behavioral", "HR Round", "Offer", "Rejected", "Withdrawn", "No Response"])
    add_validation(ws, "K", 5, end_row, ["Yes", "No"])
    add_validation(ws, "L", 5, end_row, ["—", "OA", "Phone Screen", "Technical 1", "Technical 2", "System Design", "Behavioral", "HR", "Final", "Offer"])
    add_validation(ws, "O", 5, end_row, ["Yes", "No"])

    # Column widths
    widths = [4, 22, 42, 16, 12, 8, 10, 14, 14, 16, 10, 16, 14, 14, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


def build_recruiter_sheet(wb):
    ws = wb.create_sheet("Recruiter Outreach")

    headers = [
        "#",
        "Name",
        "Company",
        "Role/Title",
        "Channel",
        "Date Contacted",
        "Message Sent",
        "Response",
        "Response Date",
        "Follow-up Date",
        "Status",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers), C_HEADER2)

    for row in range(2, 22):
        ws.cell(row=row, column=1, value=row - 1)
        style_row(ws, row, len(headers), alt=(row % 2 == 0))

    add_validation(ws, "E", 2, 21, ["LinkedIn InMail", "LinkedIn Connection", "Email", "Naukri", "Foundit", "Other"])
    add_validation(ws, "H", 2, 21, ["No Response", "Viewed", "Replied", "Call Scheduled", "Referred", "Declined"])
    add_validation(ws, "K", 2, 21, ["Pending", "In Progress", "Completed", "No Response", "Closed"])

    widths = [4, 22, 20, 30, 18, 14, 14, 14, 14, 14, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def build_referral_sheet(wb):
    ws = wb.create_sheet("Referrals")

    headers = [
        "#",
        "Referrer Name",
        "Company",
        "Role",
        "Connection",
        "Date Requested",
        "Referral Given",
        "Application Status",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers), C_HEADER4)

    for row in range(2, 17):
        ws.cell(row=row, column=1, value=row - 1)
        style_row(ws, row, len(headers), alt=(row % 2 == 0))

    add_validation(ws, "E", 2, 16, ["1st", "2nd", "3rd", "Alumni", "Cold"])
    add_validation(ws, "G", 2, 16, ["Yes", "No", "Pending"])

    widths = [4, 22, 20, 35, 12, 14, 14, 16, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")

    ws.merge_cells("A1:D1")
    ws["A1"].value = "JOB SEARCH DASHBOARD"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws["A2"].value = f"Updated: {datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(size=10, color="666666")

    metrics = [
        ("Total Applications", "=COUNTA(Applications!B5:B100)"),
        ("S-tier Applied", '=COUNTIF(Applications!F5:F100,"S")'),
        ("A-tier Applied", '=COUNTIF(Applications!F5:F100,"A")'),
        ("B-tier Applied", '=COUNTIF(Applications!F5:F100,"B")'),
        ("Callbacks Received", '=COUNTIF(Applications!K5:K100,"Yes")'),
        ("OA/Interviews", '=COUNTIFS(Applications!J5:J100,"<>Applied",Applications!J5:J100,"<>Not Applied",Applications!J5:J100,"<>No Response",Applications!J5:J100,"<>")'),
        ("Offers", '=COUNTIF(Applications!J5:J100,"Offer")'),
        ("Rejections", '=COUNTIF(Applications!J5:J100,"Rejected")'),
        ("No Response", '=COUNTIF(Applications!J5:J100,"No Response")'),
        ("Pending Follow-up", '=COUNTIF(Applications!O5:O100,"No")'),
        ("Recruiter Contacts", "=COUNTA('Recruiter Outreach'!B2:B100)"),
        ("Referrals Requested", "=COUNTA(Referrals!B2:B100)"),
    ]

    headers = ["Metric", "Count"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 2, C_HEADER3)

    for i, (label, formula) in enumerate(metrics, 5):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)
        ws.cell(row=i, column=1).font = Font(bold=True, size=11)
        ws.cell(row=i, column=2).font = Font(size=14, bold=True, color=C_HEADER)
        ws.cell(row=i, column=2).alignment = CENTER
        for col in range(1, 3):
            ws.cell(row=i, column=col).border = BORDER

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14

    # Country breakdown
    ws.cell(row=4, column=4, value="Country")
    ws.cell(row=4, column=5, value="Applications")
    style_header(ws, 4, 5, C_HEADER2)

    countries = ["Ireland", "Netherlands", "Singapore", "Indonesia", "UAE", "India", "Germany", "Sweden", "Switzerland", "UK"]
    for i, country in enumerate(countries, 5):
        ws.cell(row=i, column=4, value=country)
        ws.cell(row=i, column=5, value=f'=COUNTIF(Applications!E5:E100,"{country}")')
        ws.cell(row=i, column=4).font = BODY_FONT
        ws.cell(row=i, column=5).font = Font(size=12, bold=True, color=C_HEADER2)
        ws.cell(row=i, column=5).alignment = CENTER
        for col in range(4, 6):
            ws.cell(row=i, column=col).border = BORDER

    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    ws.sheet_properties.tabColor = C_HEADER3

    # Move Dashboard to first position
    wb.move_sheet(ws, offset=-3)


def build():
    wb = Workbook()
    build_applications_sheet(wb)
    build_recruiter_sheet(wb)
    build_referral_sheet(wb)
    build_dashboard_sheet(wb)
    wb.save(str(TRACKER_FILE))
    print(f"Built application tracker: {TRACKER_FILE}")


if __name__ == "__main__":
    build()
